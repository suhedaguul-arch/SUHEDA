import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

# desc,colour,type,pcs/box,box_vol,box/layer,layer/pallet,weight_gr,price,family,price_known
P=[
 ("PREMIUM FORK","TRANSPARENT","PALET",4000,0.046,7,10,2.8,7.5,"PREMIUM",True),
 ("PREMIUM FORK","BLACK","PALET",4000,0.046,7,10,2.8,7.5,"PREMIUM",True),
 ("PREMIUM SPOON","TRANSPARENT","PALET",4000,0.056,7,8,2.8,7.5,"PREMIUM",True),
 ("PREMIUM SPOON","BLACK","PALET",4000,0.056,7,8,2.8,7.5,"PREMIUM",True),
 ("PREMIUM KNIFE","TRANSPARENT","PALET",4000,0.035,7,11,2.8,7.5,"PREMIUM",True),
 ("PREMIUM KNIFE","BLACK","PALET",4000,0.035,7,11,2.8,7.5,"PREMIUM",True),
 ("PREMIUM DESSERT SPOON","TRANSPARENT","PALET",4000,0.036,7,9,2.5,7.25,"PREMIUM",True),
 ("PREMIUM DESSERT SPOON","BLACK","PALET",4000,0.036,7,9,2.5,7.25,"PREMIUM",True),
 ("DIAMOND FORK","TRANSPARENT","PALET",2000,0.024,12,10,3.8,10.0,"DIAMOND",True),
 ("DIAMOND FORK","BLACK","PALET",2000,0.024,12,10,3.8,10.0,"DIAMOND",True),
 ("DIAMOND SPOON","TRANSPARENT","PALET",2000,0.031,12,8,3.8,10.0,"DIAMOND",True),
 ("DIAMOND SPOON","BLACK","PALET",2000,0.031,12,8,3.8,10.0,"DIAMOND",True),
 ("DIAMOND KNIFE","TRANSPARENT","PALET",2000,0.020,12,12,3.8,10.0,"DIAMOND",True),
 ("DIAMOND KNIFE","BLACK","PALET",2000,0.020,12,12,3.8,10.0,"DIAMOND",True),
 ("SMART FORK","TRANSPARENT / BLACK","PALET",2000,0.025,12,10,3.7,9.75,"SMART",True),
 ("WAVY ICE CREAM SPOON","COLORED (RED/BLUE/GREEN/ORANGE)","PALET",2000,0.025,9,9,2.4,8.5,"ICE CREAM",True),
 ("120 CC CHAMPAGNE GLASS","TRANSPARENT (FOOT COLORED)","PALET",300,0.053,6,6,12.5,100.0,"GLASS",True),
 ("170 CC MINI WINE GLASS","TRANSPARENT (FOOT COLORED)","PALET",300,0.049,6,10,12.5,100.0,"GLASS",True),
 ("185 CC WINE GLASS","TRANSPARENT (FOOT COLORED)","PALET",300,0.067,6,7,14.5,110.0,"GLASS",True),
 ("PIZZA TRIPOD","WHITE","PALET",1000,0.027,9,11,1.85,8.5,"PIZZA",True),
 # ---- extra products (NAVIPACK volume 07.07.2026, not in price offer -> price to be entered) ----
 ("PANDA ICE CREAM SPOON","WHITE","PALET",8000,0.047,7,10,0.9,None,"ICE CREAM",False),
 ("SHOT GLASS","TRANSPARENT","DOKME",2000,0.035,None,None,3.2,None,"GLASS",False),
 ("95 MM FLAT ICE CREAM SPOON","COLORED (YELLOW/RED/BLACK)","DOKME",18000,0.083,None,None,1.25,None,"ICE CREAM",False),
]
TRUCK_DEFAULT=90

BRAND="1F3864"; HEADER="305496"; INPUTBG="FFF2CC"; INPUTHD="BF8F00"
OUTBG="E2EFDA"; OUTHD="548235"; REFBG="F2F2F2"; REFHD="A6A6A6"; KPIBG="DDEBF7"
DOKMEBG="FCE4D6"; WHITE="FFFFFF"
thin=Side(style="thin",color="BFBFBF"); med=Side(style="medium",color="808080")
b_all=Border(left=thin,right=thin,top=thin,bottom=thin); b_box=Border(left=med,right=med,top=med,bottom=med)
def Hd(c,fill=HEADER,color=WHITE,size=10):
    c.font=Font(bold=True,color=color,size=size); c.fill=PatternFill("solid",fgColor=fill)
    c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); c.border=b_all
def money(c):c.number_format='#,##0.00'
def num0(c):c.number_format='#,##0'
def vol(c):c.number_format='0.000'
def wt(c):c.number_format='#,##0.0'
def pal(c):c.number_format='0.00'
def pct(c):c.number_format='0.0%'

wb=openpyxl.Workbook()
ws=wb.active; ws.title="SIPARIS - ORDER"; ws.sheet_view.showGridLines=False

ws.merge_cells("A1:E4")
t=ws["A1"]; t.value="NAVIPACK LLC\nSIPARIS & HACIM HESAPLAMA\nORDER & VOLUME CALCULATOR (v3)"
t.font=Font(bold=True,color=WHITE,size=13); t.fill=PatternFill("solid",fgColor=BRAND)
t.alignment=Alignment(horizontal="left",vertical="center",wrap_text=True)
def lbl(cell,txt):
    ws[cell]=txt; ws[cell].font=Font(bold=True,size=9); ws[cell].alignment=Alignment(horizontal="right",vertical="center")
def inp(cell,val,fmt=None):
    c=ws[cell]; c.value=val; c.fill=PatternFill("solid",fgColor=INPUTBG); c.border=b_box
    c.alignment=Alignment(horizontal="center",vertical="center"); c.font=Font(bold=True)
    if fmt:fmt(c)
    return c
lbl("F1","MUSTERI / CUSTOMER:"); ws.merge_cells("G1:I1"); inp("G1","")
lbl("F2","TARIH / DATE:"); inp("G2","=TODAY()"); ws["G2"].number_format="dd.mm.yyyy"
lbl("F3","TIR KAPASITESI (m3):"); inp("G3",TRUCK_DEFAULT)
_tc=Comment("NAVIPACK standart TIR = 90 m3\nCift dorse (double trailer) = 120 m3\nDolum HACME gore hesaplanir.","NAVIPACK")
_tc.width=220;_tc.height=80; ws["G3"].comment=_tc

HR=6
cols=[("A","NO",4),("B","URUN / PRODUCT",24),("C","RENK / COLOUR",20),("D","YUKLEME\nTIPI",10),
 ("E","ADET/KOLI\n(pcs/box)",10),("F","KOLI HACMI\nm3",9),("G","KOLI NET\nkg",8),
 ("H","KOLI/SIRA\n(box/layer)",10),("I","SIRA/PALET\n(layer/pal)",10),("J","KOLI/PALET\n(box/pal)",9),
 ("K","FIYAT/1000\nUSD",10),("L","SIPARIS\n(KOLI) ***",12),
 ("M","SIPARIS\n(ADET)",11),("N","PALET\n(esdeger)",9),("O","TAM SIRA\n(rows)",8),
 ("P","NET AGIRLIK\nkg",11),("Q","HACIM\nm3",9),("R","TUTAR\nUSD",12),("S","AILE",12)]
for col,txt,w in cols:
    ws.column_dimensions[col].width=w; c=ws[f"{col}{HR}"]; c.value=txt
    if col=="D": Hd(c,fill="833C00")
    elif col=="L": Hd(c,fill=INPUTHD)
    elif col in "MNOPQR": Hd(c,fill=OUTHD)
    elif col=="S": Hd(c,fill=REFHD,size=9)
    elif col=="K": Hd(c,fill=INPUTHD)
    else: Hd(c)
ws.row_dimensions[HR].height=30

first=HR+1
for i,row in enumerate(P):
    r=first+i
    desc,colour,typ,pcsbox,bvol,bpl,lpp,wgr,price,fam,pk=row
    ws[f"A{r}"]=i+1; ws[f"B{r}"]=desc; ws[f"C{r}"]=colour; ws[f"D{r}"]=typ
    ws[f"E{r}"]=pcsbox; ws[f"F{r}"]=bvol; ws[f"G{r}"]=f"=E{r}*{wgr}/1000"
    if bpl: ws[f"H{r}"]=bpl
    if lpp: ws[f"I{r}"]=lpp
    ws[f"J{r}"]=f"=IF(OR(H{r}=\"\",I{r}=\"\"),0,H{r}*I{r})"
    if pk: ws[f"K{r}"]=price
    else:  ws[f"K{r}"]=None  # user enters price
    ws[f"L{r}"]=0
    ws[f"M{r}"]=f"=L{r}*E{r}"
    ws[f"N{r}"]=f'=IF(D{r}="DOKME",0,IF(J{r}=0,0,L{r}/J{r}))'
    ws[f"O{r}"]=f'=IF(D{r}="DOKME",0,IF(H{r}=0,0,L{r}/H{r}))'
    ws[f"P{r}"]=f"=L{r}*G{r}"
    ws[f"Q{r}"]=f"=L{r}*F{r}"
    ws[f"R{r}"]=f"=IF(K{r}=\"\",0,M{r}/1000*K{r})"
    ws[f"S{r}"]=fam
    num0(ws[f"E{r}"]); vol(ws[f"F{r}"]); wt(ws[f"G{r}"]); money(ws[f"K{r}"])
    num0(ws[f"L{r}"]); num0(ws[f"M{r}"]); pal(ws[f"N{r}"]); ws[f"O{r}"].number_format='0'
    wt(ws[f"P{r}"]); vol(ws[f"Q{r}"]); money(ws[f"R{r}"])
    for col in "ABCDEFGHIJK": ws[f"{col}{r}"].border=b_all
    ws[f"S{r}"].border=b_all; ws[f"S{r}"].fill=PatternFill("solid",fgColor=REFBG)
    # input cells: L always; K yellow only when price unknown
    kc=ws[f"L{r}"]; kc.fill=PatternFill("solid",fgColor=INPUTBG); kc.border=b_box; kc.font=Font(bold=True); kc.alignment=Alignment(horizontal="center")
    if not pk:
        pc=ws[f"K{r}"]; pc.fill=PatternFill("solid",fgColor=INPUTBG); pc.border=b_box; pc.alignment=Alignment(horizontal="center")
    for col in "MNOPQR": ws[f"{col}{r}"].fill=PatternFill("solid",fgColor=OUTBG); ws[f"{col}{r}"].border=b_all
    ws[f"B{r}"].font=Font(bold=True); ws[f"A{r}"].alignment=Alignment(horizontal="center")
    ws[f"C{r}"].font=Font(size=9); ws[f"C{r}"].alignment=Alignment(wrap_text=True,vertical="center")
    for col in "DEFGHIJK": ws[f"{col}{r}"].alignment=Alignment(horizontal="center")
    # highlight DOKME rows type cell
    dcell=ws[f"D{r}"]
    if typ=="DOKME":
        dcell.fill=PatternFill("solid",fgColor=DOKMEBG); dcell.font=Font(bold=True,color="833C00")
    else:
        dcell.font=Font(size=9,color="833C00")
    if not pk:
        cm=Comment("Bu urun NAVIPACK fiyat teklifinde YOK. FIYAT/1000 hucresini (K) elle giriniz.","NAVIPACK")
        cm.width=240;cm.height=70; ws[f"K{r}"].comment=cm

last=first+len(P)-1; tr=last+1
ws.merge_cells(f"A{tr}:K{tr}")
a=ws[f"A{tr}"]; a.value="TOPLAM / GRAND TOTAL"; a.font=Font(bold=True,color=WHITE,size=11)
a.fill=PatternFill("solid",fgColor=BRAND); a.alignment=Alignment(horizontal="right",vertical="center")
for col,fn in [("L",num0),("M",num0),("N",pal),("P",wt),("Q",vol),("R",money)]:
    ws[f"{col}{tr}"]=f"=SUM({col}{first}:{col}{last})"; fn(ws[f"{col}{tr}"])
for col in "LMNOPQR":
    c=ws[f"{col}{tr}"]; c.font=Font(bold=True); c.fill=PatternFill("solid",fgColor="D9E1F2"); c.border=b_box
ws.row_dimensions[tr].height=20

def kpi(anchor,label,formula,kind):
    r=int(anchor[1:]); col=anchor[0]
    ws.merge_cells(f"{col}{r}:{chr(ord(col)+1)}{r}")
    lc=ws[anchor]; lc.value=label; lc.font=Font(bold=True,size=9,color=BRAND); lc.fill=PatternFill("solid",fgColor=KPIBG)
    lc.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); lc.border=b_all
    vcol=chr(ord(col)+2); ws.merge_cells(f"{vcol}{r}:{chr(ord(col)+3)}{r}")
    vc=ws[f"{vcol}{r}"]; vc.value=formula.format(tr=tr); vc.font=Font(bold=True,size=12,color="C00000")
    vc.alignment=Alignment(horizontal="center",vertical="center"); vc.fill=PatternFill("solid",fgColor=WHITE); vc.border=b_box
    {"vol":vol,"pct":pct,"money":money}.get(kind,lambda c:setattr(c,"number_format","0"))(vc)
kpi("L1","TOPLAM\nHACIM m3","=Q{tr}","vol")
kpi("L3","GEREKEN\nTIR ADEDI","=IF($G$3=0,0,CEILING(Q{tr}/$G$3,1))","int")
kpi("P1","TIR\nDOLULUK %","=IF($G$3=0,0,Q{tr}/$G$3)","pct")
kpi("P3","TOPLAM\nTUTAR USD","=R{tr}","money")
ws.row_dimensions[1].height=26; ws.row_dimensions[3].height=26

nr=tr+1
ws.merge_cells(f"A{nr}:S{nr}")
ws[f"A{nr}"]=("*** PALET urunlerinde SIPARIS (KOLI), KOLI/SIRA (H) degerinin KATI olmalidir (1 palet sirasi = 1 tam sira; "
 "or. Diamond Fork 12/sira -> 12,24,36,48...). DOKME (palet ustu dokme / bulk) urunlerinde her adet girilebilir, palet olusturmaz, "
 "sadece HACME katkida bulunur. Fiyati bilinmeyen ekstra urunlerde FIYAT/1000 hucresi saridir - elle giriniz.")
ws[f"A{nr}"].font=Font(italic=True,size=9,color="C00000"); ws[f"A{nr}"].alignment=Alignment(wrap_text=True,vertical="center")
ws.row_dimensions[nr].height=46

# conditional validation (PALET rows: multiple of box/layer; DOKME: any qty)
dv=DataValidation(type="custom",
    formula1=f'=IF(D{first}="DOKME",L{first}>=0,AND(L{first}>=0,MOD(L{first},H{first})=0))',
    allow_blank=True,showInputMessage=True,showErrorMessage=True)
dv.promptTitle="SIPARIS (KOLI)"; dv.prompt=("PALET urun: KOLI cinsinden, KOLI/SIRA (H) KATI girin (or. 12,24,36,48).\n"
 "DOKME urun: istediginiz adet - palet olusturmaz, sadece hacme katilir.")
dv.errorTitle="GECERSIZ ADET"; dv.error="PALET urununde sayi KOLI/SIRA (H) katindan olmali (or. 12'nin katlari)."
ws.add_data_validation(dv); dv.sqref=f"L{first}:L{last}"
# loading type dropdown
dvt=DataValidation(type="list",formula1='"PALET,DOKME"',allow_blank=False)
ws.add_data_validation(dvt); dvt.sqref=f"D{first}:D{last}"
kc=Comment("Siparisi KOLI cinsinden girin. PALET urunlerde sadece KOLI/SIRA katlari kabul edilir.","NAVIPACK")
kc.width=260;kc.height=90; ws[f"L{HR}"].comment=kc
dtc=Comment("PALET = paletli yuklenir (tam sira zorunlu). DOKME = palet ustu dokme / bulk (serbest adet).","NAVIPACK")
dtc.width=240;dtc.height=70; ws[f"D{HR}"].comment=dtc

ws.freeze_panes=f"A{first}"

# ---------------- SHEET 2 : PALLET & TRUCK ----------------
ps=wb.create_sheet("PALET & TIR OZETI"); ps.sheet_view.showGridLines=False
ps.merge_cells("A1:G2"); ps["A1"]="PALET & TIR OZETI / PALLET & TRUCK SUMMARY"
ps["A1"].font=Font(bold=True,color=WHITE,size=13); ps["A1"].fill=PatternFill("solid",fgColor=BRAND); ps["A1"].alignment=Alignment(horizontal="left",vertical="center")
ps.merge_cells("A3:G4")
ps["A3"]=("Yarim paletler ONCELIKLE ayni urun ailesi icinde birlestirilir (Diamond'lar Diamond'larla, Premium'lar Premium'larla). "
 "DOKME urunler palet olusturmaz; 'palet ustu dokme' olarak mevcut paletlerin ustune yuklenir, sadece hacme katkida bulunur. "
 "'GERCEK PALET' = her ailenin palet-esdegerinin yukari yuvarlanmisi.")
ps["A3"].font=Font(italic=True,size=9,color="404040"); ps["A3"].alignment=Alignment(wrap_text=True,vertical="center")
oref="'SIPARIS - ORDER'"
ph=6
phdr=["AILE / FAMILY","SIPARIS KOLI","PALET ESDEGER","TAM PALET","GERCEK PALET","NET AGIRLIK kg","HACIM m3"]
pw=[22,13,14,11,15,14,12]
for j,(txt,w) in enumerate(zip(phdr,pw)):
    col=get_column_letter(j+1); ps.column_dimensions[col].width=w; Hd(ps[f"{col}{ph}"])
ps.row_dimensions[ph].height=30
fams=["DIAMOND","PREMIUM","SMART","ICE CREAM","GLASS","PIZZA"]
flabel={"DIAMOND":"DIAMOND CUTLERY","PREMIUM":"PREMIUM CUTLERY","SMART":"SMART FORK","ICE CREAM":"ICE CREAM SPOON","GLASS":"GLASSES / CUPS","PIZZA":"PIZZA TRIPOD"}
p0=ph+1
for k,fam in enumerate(fams):
    r=p0+k
    ps[f"A{r}"]=flabel[fam]; ps[f"A{r}"].font=Font(bold=True); ps[f"I{r}"]=fam
    # PALET rows only for boxes/pallet; but volume/weight include all of family
    ps[f"B{r}"]=f'=SUMIFS({oref}!$L${first}:$L${last},{oref}!$S${first}:$S${last},$I{r},{oref}!$D${first}:$D${last},"PALET")'
    ps[f"C{r}"]=f"=SUMIF({oref}!$S${first}:$S${last},$I{r},{oref}!$N${first}:$N${last})"
    ps[f"D{r}"]=f"=INT(C{r})"; ps[f"E{r}"]=f"=IF(C{r}=0,0,CEILING(C{r},1))"
    ps[f"F{r}"]=f"=SUMIF({oref}!$S${first}:$S${last},$I{r},{oref}!$P${first}:$P${last})"
    ps[f"G{r}"]=f"=SUMIF({oref}!$S${first}:$S${last},$I{r},{oref}!$Q${first}:$Q${last})"
    num0(ps[f"B{r}"]); pal(ps[f"C{r}"]); ps[f"D{r}"].number_format='0'; ps[f"E{r}"].number_format='0'; wt(ps[f"F{r}"]); vol(ps[f"G{r}"])
    for col in "ABCDEFG":
        cc=ps[f"{col}{r}"]; cc.border=b_all
        if col=="E": cc.fill=PatternFill("solid",fgColor=OUTBG); cc.font=Font(bold=True)
ps.column_dimensions["I"].hidden=True
# DOKME summary line
dr=p0+len(fams)
ps[f"A{dr}"]="PALET USTU DOKME / BULK"; ps[f"A{dr}"].font=Font(bold=True,color="833C00")
ps[f"A{dr}"].fill=PatternFill("solid",fgColor=DOKMEBG)
ps[f"B{dr}"]=f'=SUMIF({oref}!$D${first}:$D${last},"DOKME",{oref}!$L${first}:$L${last})'
ps[f"C{dr}"]=0; ps[f"D{dr}"]=0; ps[f"E{dr}"]=0
ps[f"F{dr}"]=f'=SUMIF({oref}!$D${first}:$D${last},"DOKME",{oref}!$P${first}:$P${last})'
ps[f"G{dr}"]=f'=SUMIF({oref}!$D${first}:$D${last},"DOKME",{oref}!$Q${first}:$Q${last})'
num0(ps[f"B{dr}"]); pal(ps[f"C{dr}"]); wt(ps[f"F{dr}"]); vol(ps[f"G{dr}"])
for col in "ABCDEFG":
    ps[f"{col}{dr}"].border=b_all
    if col!="A": ps[f"{col}{dr}"].fill=PatternFill("solid",fgColor="FBE9DD")
pt=dr+1
ps[f"A{pt}"]="TOPLAM / GRAND TOTAL"; ps[f"A{pt}"].font=Font(bold=True,color=WHITE)
ps[f"A{pt}"].fill=PatternFill("solid",fgColor=BRAND); ps[f"A{pt}"].alignment=Alignment(horizontal="right")
for col in "BCDEFG": ps[f"{col}{pt}"]=f"=SUM({col}{p0}:{col}{pt-1})"
num0(ps[f"B{pt}"]); pal(ps[f"C{pt}"]); ps[f"D{pt}"].number_format='0'; ps[f"E{pt}"].number_format='0'; wt(ps[f"F{pt}"]); vol(ps[f"G{pt}"])
for col in "BCDEFG":
    c=ps[f"{col}{pt}"]; c.font=Font(bold=True); c.fill=PatternFill("solid",fgColor="D9E1F2"); c.border=b_box

tb=pt+2
ps.merge_cells(f"A{tb}:G{tb}"); ps[f"A{tb}"]="TIR / TRUCK YUKLEME (hacme gore / by volume)"
ps[f"A{tb}"].font=Font(bold=True,color=WHITE,size=11); ps[f"A{tb}"].fill=PatternFill("solid",fgColor="833C00")
ps[f"A{tb}"].alignment=Alignment(vertical="center"); ps.row_dimensions[tb].height=20
def trow(r,label,formula,kind,hl=False):
    ps.merge_cells(f"A{r}:D{r}"); ps[f"A{r}"]=label; ps[f"A{r}"].font=Font(bold=True,size=10); ps[f"A{r}"].alignment=Alignment(horizontal="right",vertical="center")
    ps.merge_cells(f"E{r}:G{r}"); c=ps[f"E{r}"]; c.value=formula; c.alignment=Alignment(horizontal="center",vertical="center")
    c.font=Font(bold=True,size=12,color="C00000" if hl else "000000"); c.border=b_box
    if hl: c.fill=PatternFill("solid",fgColor=DOKMEBG)
    {"vol":vol,"pct":pct,"m":money}.get(kind,lambda x:setattr(x,"number_format","0"))(c)
    for col in "ABCD": ps[f"{col}{r}"].border=b_all
trow(tb+1,"TIR KAPASITESI (m3):","='SIPARIS - ORDER'!$G$3","vol")
trow(tb+2,"TOPLAM SIPARIS HACMI (m3):",f"=G{pt}","vol")
trow(tb+3,"TIR DOLULUK ORANI:",f"=IF(E{tb+1}=0,0,G{pt}/E{tb+1})","pct")
trow(tb+4,"GEREKEN TIR ADEDI:",f"=IF(E{tb+1}=0,0,CEILING(G{pt}/E{tb+1},1))","int",hl=True)
trow(tb+5,"SON TIR'DA BOS HACIM (m3):",f"=IF(E{tb+1}=0,0,E{tb+4}*E{tb+1}-G{pt})","vol")
note=tb+7
ps.merge_cells(f"A{note}:G{note+3}")
ps[f"A{note}"]=("NOT: TIR dolulugu HACME gore hesaplanir (NAVIPACK standart = 90 m3; cift dorse = 120 m3). "
 "DOKME urunlerin hacmi toplama dahildir ancak palet sayisina eklenmez (mevcut paletlerin ustune yuklenir). "
 "Farkli aileler yalnizca koli yukseklikleri uyumluysa ayni palette birlestirilebilir. Nihai istifleme fabrikada teyit edilir.")
ps[f"A{note}"].font=Font(italic=True,size=9,color="404040"); ps[f"A{note}"].alignment=Alignment(wrap_text=True,vertical="top")

# ---------------- SHEET 3 : NOTES ----------------
nt=wb.create_sheet("ACIKLAMA - NOTES"); nt.sheet_view.showGridLines=False
nt.column_dimensions["A"].width=3; nt.column_dimensions["B"].width=112
lines=[
 ("NAVIPACK - SIPARIS & HACIM HESAPLAMA / ORDER & VOLUME CALCULATOR (v3)","title"),("",""),
 ("NASIL KULLANILIR / HOW TO USE","h"),
 ("1) 'SIPARIS - ORDER' sayfasinda sari 'SIPARIS (KOLI)' sutununu (L) doldurun.","n"),
 ("   Ustten Musteri, Tarih ve TIR Kapasitesini (varsayilan 90 m3) girin.","n"),
 ("2) Ust bantta KPI: Toplam Hacim, TIR Doluluk %, Gereken TIR adedi, Toplam Tutar.","n"),
 ("3) 'PALET & TIR OZETI' paletleri aile bazinda birlestirir + TIR dolulugunu gosterir.","n"),("",""),
 ("YUKLEME TIPI: PALET vs DOKME","h"),
 ("PALET = paletli yuklenir. SIPARIS, KOLI/SIRA (H) degerinin KATI olmalidir (tam sira).","n"),
 ("DOKME = 'palet ustu dokme' / bulk. Serbest adet girilir; palet olusturmaz, sadece HACME","n"),
 ("   katkida bulunur (mevcut paletlerin ustune / TIR bosluguna yuklenir). Or. 95mm Flat Kasik, Shot Glass.","n"),("",""),
 ("NEDEN KOLI/SIRA KATLARI?","h"),
 ("Bir palet sirasina farkli urun dizilirse koli yukseklikleri farkli olur, palet dengesiz yukselir.","n"),
 ("Bu yuzden her sira TEK urunle DOLU olmali; siparis tam sira (KOLI/SIRA kati) girilir.","n"),
 ("Ornek: Diamond Fork 12/sira -> 12, 24, 36, 48 ... (45 yerine 48).","n"),("",""),
 ("TIR / TRUCK","h"),
 ("NAVIPACK standart TIR = 90 m3 (dolum HACME gore). Cift dorse TIR = 120 m3.","n"),
 ("FERPROM gibi TAMAMEN dokme yukleyen firmalarda palet yoktur; ayri 'FERPROM Dokme' dosyasina bakin.","n"),("",""),
 ("EKSTRA URUNLER (teklifte fiyat yok)","h"),
 ("PANDA / 95mm FLAT / SHOT GLASS: NAVIPACK volume tablosunda var, fiyat teklifinde yok.","n"),
 ("Bunlarin FIYAT/1000 hucresi sari - siparise gore elle giriniz.","n"),("",""),
 ("HESAPLAMALAR","h"),
 ("SIPARIS (ADET) = Koli x Adet/Koli | PALET (esdeger)=Koli/(Koli/Palet) | TAM SIRA=Koli/(Koli/Sira)","n"),
 ("NET AGIRLIK=Koli x (Adet/Koli x gr /1000) | HACIM=Koli x Koli Hacmi | TUTAR=(Adet/1000) x Fiyat","n"),("",""),
 ("RENK KODLARI","h"),
 ("Sari=veri girisi | Yesil=otomatik sonuc | Turuncu=DOKME/bulk | Gri=referans | Mavi=KPI.","n"),("",""),
 ("KAYNAKLAR","h"),
 ("NAVIPACK Price Offer 20.06.2026 + Volume Calculation 07.07.2026. TIR: standart 90 m3.","n"),
]
r=1
for txt,kind in lines:
    c=nt[f"B{r}"]; c.value=txt
    if kind=="title": c.font=Font(bold=True,color=WHITE,size=13); c.fill=PatternFill("solid",fgColor=BRAND); nt.row_dimensions[r].height=24; c.alignment=Alignment(vertical="center")
    elif kind=="h": c.font=Font(bold=True,color=BRAND,size=11); c.fill=PatternFill("solid",fgColor="D9E1F2")
    else: c.font=Font(size=10)
    r+=1

wb.active=wb.sheetnames.index("SIPARIS - ORDER")
out="NAVIPACK_Siparis_Hacim_Hesaplama.xlsx"; wb.save(out); print("SAVED",out,"| rows",first,"-",last,"| total",tr)
