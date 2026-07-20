import openpyxl, json
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

M=json.load(open("ferprom_master.json"))   # [no,desc,colour,pcs/box,vol,price1000,wt_gr]
TRUCK_DEFAULT=120   # FERPROM cift dorse

BRAND="4A1C1C"; HEADER="843C3C"; INPUTBG="FFF2CC"; INPUTHD="BF8F00"
OUTBG="E2EFDA"; OUTHD="548235"; KPIBG="F2DCDB"; WHITE="FFFFFF"
thin=Side(style="thin",color="BFBFBF"); med=Side(style="medium",color="808080")
b_all=Border(left=thin,right=thin,top=thin,bottom=thin); b_box=Border(left=med,right=med,top=med,bottom=med)
def Hd(c,fill=HEADER,color=WHITE,size=10):
    c.font=Font(bold=True,color=color,size=size); c.fill=PatternFill("solid",fgColor=fill)
    c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); c.border=b_all
def money(c):c.number_format='#,##0.00'
def num0(c):c.number_format='#,##0'
def vol(c):c.number_format='0.000'
def wt(c):c.number_format='#,##0.0'
def pct(c):c.number_format='0.0%'

wb=openpyxl.Workbook(); ws=wb.active; ws.title="SIPARIS - ORDER"; ws.sheet_view.showGridLines=False
ws.merge_cells("A1:D4")
t=ws["A1"]; t.value="FERPROM D.O.O.\nSIPARIS & HACIM HESAPLAMA (DOKME)\nBULK ORDER & VOLUME CALCULATOR"
t.font=Font(bold=True,color=WHITE,size=12); t.fill=PatternFill("solid",fgColor=BRAND); t.alignment=Alignment(horizontal="left",vertical="center",wrap_text=True)
def lbl(cell,txt):
    ws[cell]=txt; ws[cell].font=Font(bold=True,size=9); ws[cell].alignment=Alignment(horizontal="right",vertical="center")
def inp(cell,val,fmt=None):
    c=ws[cell]; c.value=val; c.fill=PatternFill("solid",fgColor=INPUTBG); c.border=b_box; c.alignment=Alignment(horizontal="center",vertical="center"); c.font=Font(bold=True)
    if fmt:fmt(c)
    return c
lbl("E1","MUSTERI / CUSTOMER:"); ws.merge_cells("F1:G1"); inp("F1","FERPROM D.O.O.")
lbl("E2","TARIH / DATE:"); inp("F2","=TODAY()"); ws["F2"].number_format="dd.mm.yyyy"
lbl("E3","TIR KAPASITESI (m3):"); inp("F3",TRUCK_DEFAULT)
_tc=Comment("FERPROM cift dorse (double trailer) = 120 m3\nStandart TIR = 90 m3\nDokme yukleme: dolum HACME gore.","SMARTPACK")
_tc.width=230;_tc.height=80; ws["F3"].comment=_tc

HR=6
cols=[("A","NO",4),("B","URUN / PRODUCT",34),("C","RENK / COLOUR",20),("D","ADET/KOLI\n(pcs/box)",10),
 ("E","KOLI HACMI\nm3",9),("F","KOLI NET\nkg",8),("G","FIYAT/1000\nUSD",10),
 ("H","SIPARIS\n(KOLI) ***",12),("I","SIPARIS\n(ADET)",12),("J","NET AGIRLIK\nkg",11),("K","HACIM\nm3",9),("L","TUTAR\nUSD",12)]
for col,txt,w in cols:
    ws.column_dimensions[col].width=w; c=ws[f"{col}{HR}"]; c.value=txt
    if col=="H": Hd(c,fill=INPUTHD)
    elif col in "IJKL": Hd(c,fill=OUTHD)
    else: Hd(c)
ws.row_dimensions[HR].height=30

first=HR+1
for i,row in enumerate(M):
    r=first+i; no,desc,colour,pcs,bvol,pr1000,wg=row
    ws[f"A{r}"]=i+1; ws[f"B{r}"]=desc; ws[f"C{r}"]=colour
    ws[f"D{r}"]=pcs; ws[f"E{r}"]=bvol; ws[f"F{r}"]=f"=D{r}*{wg}/1000"; ws[f"G{r}"]=pr1000
    ws[f"H{r}"]=0
    ws[f"I{r}"]=f"=H{r}*D{r}"; ws[f"J{r}"]=f"=H{r}*F{r}"; ws[f"K{r}"]=f"=H{r}*E{r}"; ws[f"L{r}"]=f"=I{r}/1000*G{r}"
    num0(ws[f"D{r}"]); vol(ws[f"E{r}"]); wt(ws[f"F{r}"]); money(ws[f"G{r}"]); num0(ws[f"H{r}"]); num0(ws[f"I{r}"]); wt(ws[f"J{r}"]); vol(ws[f"K{r}"]); money(ws[f"L{r}"])
    for col in "ABCDEFG": ws[f"{col}{r}"].border=b_all
    kc=ws[f"H{r}"]; kc.fill=PatternFill("solid",fgColor=INPUTBG); kc.border=b_box; kc.font=Font(bold=True); kc.alignment=Alignment(horizontal="center")
    for col in "IJKL": ws[f"{col}{r}"].fill=PatternFill("solid",fgColor=OUTBG); ws[f"{col}{r}"].border=b_all
    ws[f"B{r}"].font=Font(bold=True,size=9); ws[f"A{r}"].alignment=Alignment(horizontal="center")
    ws[f"C{r}"].font=Font(size=8); ws[f"C{r}"].alignment=Alignment(wrap_text=True,vertical="center")
    for col in "DEFG": ws[f"{col}{r}"].alignment=Alignment(horizontal="center")
    if i%2==1:
        for col in "ABCDEFG": ws[f"{col}{r}"].fill=PatternFill("solid",fgColor="FAF6F6")

last=first+len(M)-1; tr=last+1
ws.merge_cells(f"A{tr}:G{tr}")
a=ws[f"A{tr}"]; a.value="TOPLAM / GRAND TOTAL"; a.font=Font(bold=True,color=WHITE,size=11); a.fill=PatternFill("solid",fgColor=BRAND); a.alignment=Alignment(horizontal="right",vertical="center")
for col,fn in [("H",num0),("I",num0),("J",wt),("K",vol),("L",money)]:
    ws[f"{col}{tr}"]=f"=SUM({col}{first}:{col}{last})"; fn(ws[f"{col}{tr}"])
for col in "HIJKL":
    c=ws[f"{col}{tr}"]; c.font=Font(bold=True); c.fill=PatternFill("solid",fgColor="F2DCDB"); c.border=b_box
ws.row_dimensions[tr].height=20

def kpi(lcol,lrow,label,formula,kind):
    # vertical: label row = lrow (2 cols), value row = lrow+1
    c2=chr(ord(lcol)+1)
    ws.merge_cells(f"{lcol}{lrow}:{c2}{lrow}")
    lc=ws[f"{lcol}{lrow}"]; lc.value=label; lc.font=Font(bold=True,size=9,color=BRAND); lc.fill=PatternFill("solid",fgColor=KPIBG); lc.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); lc.border=b_all
    ws.merge_cells(f"{lcol}{lrow+1}:{c2}{lrow+1}")
    vc=ws[f"{lcol}{lrow+1}"]; vc.value=formula.format(tr=tr); vc.font=Font(bold=True,size=12,color="C00000"); vc.alignment=Alignment(horizontal="center",vertical="center"); vc.fill=PatternFill("solid",fgColor=WHITE); vc.border=b_box
    {"vol":vol,"pct":pct,"money":money}.get(kind,lambda c:setattr(c,"number_format","0"))(vc)
# 2x2 grid in cols I-J / K-L, rows 1-2 and 3-4
kpi("I",1,"TOPLAM\nHACIM m3","=K{tr}","vol")
kpi("K",1,"TIR\nDOLULUK %","=IF($F$3=0,0,K{tr}/$F$3)","pct")
kpi("I",3,"GEREKEN\nTIR ADEDI","=IF($F$3=0,0,CEILING(K{tr}/$F$3,1))","int")
kpi("K",3,"TOPLAM\nTUTAR USD","=L{tr}","money")
for rr in (1,2,3,4): ws.row_dimensions[rr].height=22

nr=tr+1; ws.merge_cells(f"A{nr}:L{nr}")
ws[f"A{nr}"]=("*** FERPROM tamamen DOKME (bulk) yukleme yapar - palet YOKTUR. Her urun icin istediginiz KOLI adedini girin; "
 "kisit yoktur. Yukleme yalnizca toplam HACMIN TIR'a sigmasi ile hesaplanir (cift dorse = 120 m3).")
ws[f"A{nr}"].font=Font(italic=True,size=9,color="C00000"); ws[f"A{nr}"].alignment=Alignment(wrap_text=True,vertical="center"); ws.row_dimensions[nr].height=34

dv=DataValidation(type="whole",operator="greaterThanOrEqual",formula1="0",allow_blank=True,showInputMessage=True,showErrorMessage=True)
dv.promptTitle="SIPARIS (KOLI)"; dv.prompt="KOLI cinsinden istediginiz adedi girin. Dokme yukleme - palet/sira kisiti yoktur."
dv.errorTitle="GECERSIZ"; dv.error="Lutfen 0 veya pozitif tam sayi girin."
ws.add_data_validation(dv); dv.sqref=f"H{first}:H{last}"
ws.freeze_panes=f"A{first}"

# ---------------- NOTES ----------------
nt=wb.create_sheet("ACIKLAMA - NOTES"); nt.sheet_view.showGridLines=False
nt.column_dimensions["A"].width=3; nt.column_dimensions["B"].width=110
lines=[
 ("FERPROM - DOKME SIPARIS & HACIM HESAPLAMA / BULK ORDER & VOLUME CALCULATOR","title"),("",""),
 ("DOKME (BULK) YUKLEME NEDIR?","h"),
 ("FERPROM tamamen dokme yukleme ister - PALET KULLANILMAZ. Koliler TIR'a dogrudan istiflenir.","n"),
 ("Bu yuzden palet/sira mantigi yoktur; tek kisit toplam HACMIN TIR'a sigmasidir.","n"),("",""),
 ("NASIL KULLANILIR / HOW TO USE","h"),
 ("1) Sari 'SIPARIS (KOLI)' sutununa (H) her urun icin istediginiz koli adedini girin (serbest).","n"),
 ("2) Ustten TIR Kapasitesini girin (FERPROM cift dorse = 120 m3; standart TIR = 90 m3).","n"),
 ("3) Ust bantta KPI: Toplam Hacim, TIR Doluluk %, Gereken TIR adedi, Toplam Tutar.","n"),("",""),
 ("HESAPLAMALAR / FORMULAS","h"),
 ("SIPARIS (ADET) = Koli x Adet/Koli","n"),
 ("NET AGIRLIK kg = Koli x (Adet/Koli x urun agirligi gr / 1000)","n"),
 ("HACIM m3       = Koli x Koli Hacmi m3","n"),
 ("TUTAR USD      = (Siparis Adet / 1000) x Fiyat/1000","n"),
 ("GEREKEN TIR    = Toplam Hacim / TIR Kapasitesi (yukari yuvarlanir)","n"),("",""),
 ("NAVIPACK ILE FARK","h"),
 ("NAVIPACK paletli yukler: koli/sira, palet ve tam-sira kisiti vardir (ayri dosya).","n"),
 ("FERPROM dokme yukler: sadece hacim onemlidir, palet yoktur.","n"),("",""),
 ("KAYNAKLAR / SOURCES","h"),
 ("FERPROM Price Offer 20.07.2026 + Proforma 26094. Cift dorse TIR = 120 m3.","n"),
]
r=1
for txt,kind in lines:
    c=nt[f"B{r}"]; c.value=txt
    if kind=="title": c.font=Font(bold=True,color=WHITE,size=13); c.fill=PatternFill("solid",fgColor=BRAND); nt.row_dimensions[r].height=24; c.alignment=Alignment(vertical="center")
    elif kind=="h": c.font=Font(bold=True,color=BRAND,size=11); c.fill=PatternFill("solid",fgColor="F2DCDB")
    else: c.font=Font(size=10)
    r+=1

wb.active=0
out="FERPROM_Dokme_Siparis_Hacim.xlsx"; wb.save(out); print("SAVED",out,"| products",len(M))
