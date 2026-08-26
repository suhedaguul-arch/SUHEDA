#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
SmartPack Export Dashboard — DATA AGENT / build pipeline
=========================================================

Amaç:  Ham fatura verisi + Stok Taksonomisi girildiğinde, ürünleri DEĞİŞMEZ
       StokKodu ile eşleştirip dashboard verisini (const D) sıfırdan derler ve
       dashboard.html içine enjekte eder.

TEMEL KURAL (asla ihlal edilmez):
       Ürün kimliği (Genel Ürün Adı, Tip, Seri, Renk) SADECE StokKodu üzerinden,
       taksonomi tablosundan gelir. Faturadaki UrunAdi/Kategori sütunları
       YOK SAYILIR. StokKodu asla değiştirilmez — join anahtarıdır.

Kullanım:
       python3 build_dashboard.py                 # varsayılan yollarla derle
       python3 build_dashboard.py --check         # sadece doğrula, yazma
       python3 build_dashboard.py --out yeni.html # farklı çıktıya yaz

Girdi dosyaları (data/ klasörü):
       ham_veri.xlsx        -> fatura satırları (fact table)
       taksonomi.xlsx       -> StokKodu master (STOK TAKSONOMİSİ sayfası)
Referans (pipeline/reference/):
       product_en.json      -> Genel Ürün Adı -> İngilizce ad
       seri_cat.json        -> Seri -> seri kategorisi (EN grup)
       country_geo.json     -> Ülke -> {lat, lon, Ulke_EN}
       exchange_rates.json  -> Döviz -> USD kuru
"""

import openpyxl, json, argparse, sys, os, statistics
from collections import defaultdict, Counter
from datetime import datetime, date

# ══════════════════════════════════════════════════════════════════
#  YAPILANDIRMA
# ══════════════════════════════════════════════════════════════════
HERE      = os.path.dirname(os.path.abspath(__file__))
ROOT      = os.path.dirname(HERE)
DATA      = os.path.join(ROOT, "data")
REF       = os.path.join(HERE, "reference")

PATH_HAM  = os.path.join(DATA, "ham_veri.xlsx")
PATH_TAX  = os.path.join(DATA, "taksonomi.xlsx")
PATH_HTML = os.path.join(ROOT, "dashboard.html")
PATH_IND  = os.path.join(DATA, "ihrac_kayitli.json")   # ihraç kayıtlı (indirekt)
DIR_DIA   = os.path.join(DATA, "dia")                   # ham DIA fatura xlsx klasörü

SHEET_HAM = "📋 Ham Veri (Power BI)"
SHEET_TAX = "STOK TAKSONOMİSİ"

# İhraç kayıtlı satışların pazar etiketi (yurtiçi aracı firmalar üzerinden)
IND_MARKET    = "İHRAÇ KAYITLI"
IND_MARKET_EN = "Export-Registered"

# Ham DIA "AYRINTILI İHRACAT FATURALARI" formatı: başlık 4. satır, veri 6.+
# Kolonlar (0-tabanlı): 4=CariKodu 5=Ünvan 6=StokKodu 7=Açıklama 8=Miktar
#                       10=BirimFiyat 11=ToplamTutar 12=SatırDövizi, 2=Tarih
DIA_COLS = {"Tarih":2,"FaturaNo":3,"CariKodu":4,"Unvan":5,"StokKodu":6,
            "Aciklama":7,"Miktar":8,"Birim":9,"BirimFiyat":10,
            "ToplamTutar":11,"Doviz":12}

# Ham veri kolon indeksleri (0-tabanlı) — başlık satırından doğrulanır
HAM_COLS = {
    "Tarih":0, "Yil":1, "Ay":2, "CariKodu":3, "Unvan":4, "Ulke":5,
    "Ulke_EN":6, "StokKodu":7, "UrunAdi":8, "Kategori":9, "Miktar":10,
    "Birim":11, "BirimFiyat":12, "ToplamTutar":13, "Doviz":14, "USD_Tutar":15,
}
# Taksonomi kolon indeksleri
TAX_COLS = {
    "StokKodu":0, "UrunAdiDIA":1, "GenelUrunAdi":2, "Hammadde":3,
    "UrunTipi":4, "Seri":5, "Renk":6, "USDFiyat":7, "Adet":8,
}

YEARS = [2023, 2024, 2025, 2026]

# ══════════════════════════════════════════════════════════════════
#  YARDIMCILAR
# ══════════════════════════════════════════════════════════════════
def load_json(p, default=None):
    if os.path.exists(p):
        return json.load(open(p, encoding="utf-8"))
    return default if default is not None else {}

def norm_stok(v):
    """StokKodu'yu tek biçimli tamsayı anahtarına çevir (asla değiştirmeden)."""
    if v is None: return None
    s = str(v).strip()
    if s == "": return None
    try:
        return int(float(s))
    except ValueError:
        return s  # sayısal olmayan kodları da destekle

def parse_date(v):
    if isinstance(v, (datetime, date)):
        return datetime(v.year, v.month, v.day)
    if v is None: return None
    s = str(v).strip()[:10]
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
        try: return datetime.strptime(s, fmt)
        except ValueError: pass
    return None

def num(v, d=0.0):
    if v is None: return d
    try: return float(str(v).replace(",", "."))
    except (ValueError, TypeError): return d

# ══════════════════════════════════════════════════════════════════
#  1) TAKSONOMİ  (StokKodu -> ürün kimliği)
# ══════════════════════════════════════════════════════════════════
def load_taxonomy(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[SHEET_TAX] if SHEET_TAX in wb.sheetnames else wb[wb.sheetnames[0]]
    tax = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        code = norm_stok(r[TAX_COLS["StokKodu"]])
        if code is None: continue
        genel = r[TAX_COLS["GenelUrunAdi"]]
        if not genel: continue
        tax[code] = {
            "GenelUrunAdi": str(genel).strip(),
            "Hammadde":  (r[TAX_COLS["Hammadde"]]  or "").strip() if r[TAX_COLS["Hammadde"]] else "",
            "UrunTipi":  (r[TAX_COLS["UrunTipi"]]  or "").strip() if r[TAX_COLS["UrunTipi"]] else "",
            "Seri":      (r[TAX_COLS["Seri"]]      or "").strip() if r[TAX_COLS["Seri"]] else "",
            "Renk":      (r[TAX_COLS["Renk"]]      or "Diğer").strip() if r[TAX_COLS["Renk"]] else "Diğer",
        }
    wb.close()
    return tax

# ══════════════════════════════════════════════════════════════════
#  2) HAM FATURA VERİSİ  (fact table)
# ══════════════════════════════════════════════════════════════════
def load_facts(path, tax, rates, report):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[SHEET_HAM] if SHEET_HAM in wb.sheetnames else wb[wb.sheetnames[0]]
    facts = []
    unknown_codes = Counter()
    rows = ws.iter_rows(min_row=2, values_only=True)
    for r in rows:
        if r is None or r[HAM_COLS["Tarih"]] is None:
            continue
        code = norm_stok(r[HAM_COLS["StokKodu"]])
        if code is None:
            continue
        ident = tax.get(code)
        if ident is None:
            unknown_codes[code] += 1
            continue  # taksonomide olmayan kodu atla (rapora düşer)

        dt   = parse_date(r[HAM_COLS["Tarih"]])
        yil  = int(num(r[HAM_COLS["Yil"]], dt.year if dt else 0))
        ay   = int(num(r[HAM_COLS["Ay"]],  dt.month if dt else 0))
        cari = norm_stok(r[HAM_COLS["CariKodu"]])
        # USD tutarı: hazır varsa kullan, yoksa döviz kuru ile hesapla
        usd  = num(r[HAM_COLS["USD_Tutar"]], None)
        if usd is None or usd == 0:
            doviz = str(r[HAM_COLS["Doviz"]] or "USD").strip().upper()
            usd = num(r[HAM_COLS["ToplamTutar"]]) * rates.get(doviz, 1.0)
        facts.append({
            "dt": dt, "yil": yil, "ay": ay,
            "cari": cari,
            "unvan": (r[HAM_COLS["Unvan"]] or "").strip(),
            "ulke":  (r[HAM_COLS["Ulke"]]  or "").strip(),
            "ulke_en": (r[HAM_COLS["Ulke_EN"]] or "").strip(),
            "stok": code,
            "genel": ident["GenelUrunAdi"],
            "tip":   ident["UrunTipi"],
            "seri":  ident["Seri"],
            "renk":  ident["Renk"],
            "miktar": num(r[HAM_COLS["Miktar"]]),
            "usd": usd,
        })
    wb.close()
    report["unknown_codes"] = dict(unknown_codes)
    return facts

# ══════════════════════════════════════════════════════════════════
#  2b) HAM DIA FATURALARI  (AYRINTILI İHRACAT FATURALARI xlsx)
# ══════════════════════════════════════════════════════════════════
def load_facts_from_dia(dia_dir, tax, cari_country, rates, report):
    """data/dia/ altındaki tüm 'AYRINTILI İHRACAT FATURALARI' xlsx dosyalarını
    okur. Ülke bilgisi cari_country ile CariKodu üzerinden eklenir."""
    facts = []
    unknown_codes = Counter()
    unknown_cari  = Counter()
    if not os.path.isdir(dia_dir):
        return facts
    files = sorted(f for f in os.listdir(dia_dir) if f.lower().endswith((".xlsx",".xls")))
    for fn in files:
        wb = openpyxl.load_workbook(os.path.join(dia_dir, fn), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        for r in ws.iter_rows(min_row=6, values_only=True):
            if r is None or len(r) <= DIA_COLS["Doviz"]: continue
            code = norm_stok(r[DIA_COLS["StokKodu"]])
            if code is None: continue
            ident = tax.get(code)
            if ident is None:
                unknown_codes[code] += 1; continue
            dt   = parse_date(r[DIA_COLS["Tarih"]])
            cari = norm_stok(r[DIA_COLS["CariKodu"]])
            cc   = cari_country.get(str(cari), {})
            if not cc: unknown_cari[cari] += 1
            doviz = str(r[DIA_COLS["Doviz"]] or "USD").strip().upper()
            usd   = num(r[DIA_COLS["ToplamTutar"]]) * rates.get(doviz, 1.0)
            facts.append({
                "dt":dt, "yil":dt.year if dt else 0, "ay":dt.month if dt else 0,
                "cari":cari, "unvan":(r[DIA_COLS["Unvan"]] or "").strip(),
                "ulke":cc.get("ulke",""), "ulke_en":cc.get("ulke_en",""),
                "stok":code, "genel":ident["GenelUrunAdi"], "tip":ident["UrunTipi"],
                "seri":ident["Seri"], "renk":ident["Renk"],
                "miktar":num(r[DIA_COLS["Miktar"]]), "usd":usd,
            })
        wb.close()
    report.setdefault("unknown_codes", {})
    for k,v in unknown_codes.items(): report["unknown_codes"][k] = report["unknown_codes"].get(k,0)+v
    report["unknown_cari"] = dict(unknown_cari)
    report["dia_files"] = files
    return facts

# ══════════════════════════════════════════════════════════════════
#  2c) İHRAÇ KAYITLI (indirekt) — PDF'ten çıkarılmış JSON
# ══════════════════════════════════════════════════════════════════
def load_facts_from_indirect(path, tax, rates, report):
    """İhraç kayıtlı satışlar (yurtiçi aracı firmalar). Pazar = İHRAÇ KAYITLI."""
    facts = []
    unknown = Counter()
    if not os.path.exists(path): return facts
    items = json.load(open(path, encoding="utf-8"))
    for it in items:
        code = norm_stok(it.get("kod"))
        if code is None or not it.get("tutar"): continue
        ident = tax.get(code)
        if ident is None:
            unknown[code] += 1; continue
        dt = parse_date(it.get("tarih"))
        usd = num(it["tutar"]) * rates.get(str(it.get("dov","EUR")).upper(), 1.0)
        facts.append({
            "dt":dt, "yil":dt.year if dt else 0, "ay":dt.month if dt else 0,
            "cari":"IND_"+(it.get("buyer","")[:20]), "unvan":it.get("buyer",""),
            "ulke":IND_MARKET, "ulke_en":IND_MARKET_EN,
            "stok":code, "genel":ident["GenelUrunAdi"], "tip":ident["UrunTipi"],
            "seri":ident["Seri"], "renk":ident["Renk"],
            "miktar":num(it.get("miktar")), "usd":usd,
        })
    report.setdefault("unknown_codes", {})
    for k,v in unknown.items(): report["unknown_codes"][k] = report["unknown_codes"].get(k,0)+v
    report["indirect_lines"] = len(facts)
    return facts

# ══════════════════════════════════════════════════════════════════
#  3) AGREGASYON -> D
# ══════════════════════════════════════════════════════════════════
def build_D(facts, tax, ref, report):
    prod_en    = ref["product_en"]
    seri_cat   = ref["seri_cat"]
    country_geo= ref["country_geo"]

    # referans tarih = en son fatura tarihi (deterministik "gün önce" için)
    max_dt = max((f["dt"] for f in facts if f["dt"]), default=datetime.now())

    # ---- ÜLKE ----
    ulke_agg = defaultdict(lambda: {
        "miktar":0.0, "usd_2023":0.0,"usd_2024":0.0,"usd_2025":0.0,"usd_2026":0.0,
        "musteriler":set(), "fatura":0, "toplam":0.0, "Ulke_EN":""})
    # ---- MÜŞTERİ ----
    cari_agg = defaultdict(lambda: {
        "ad":"", "ulke":"", "toplam":0.0, "fatura":0,
        "son":None, "ilk":None})
    # ---- ÜRÜN (GenelUrunAdi) ----
    urun_agg = defaultdict(lambda: {
        "tip":"", "seri":"", "miktar":0.0, "usd":0.0,
        "usd_2023":0.0,"usd_2024":0.0,"usd_2025":0.0,"usd_2026":0.0,
        "renkler":defaultdict(lambda:{"usd":0.0,"miktar":0.0})})
    # ---- RENK ----
    renk_agg = defaultdict(lambda: {"USD_Tutar":0.0,"miktar":0.0})
    # ---- KATEGORİ (UrunTipi) ----
    kat_agg  = defaultdict(lambda: {"USD_Tutar":0.0,"miktar":0.0})
    # ---- TREND ----
    trend_y  = defaultdict(float)
    trend_m  = defaultdict(float)

    for f in facts:
        u, c, g = f["ulke"], f["cari"], f["genel"]
        usd, mik, yil = f["usd"], f["miktar"], f["yil"]

        # Ülke
        if u:
            a = ulke_agg[u]
            a["miktar"] += mik; a["toplam"] += usd
            if yil in YEARS: a[f"usd_{yil}"] += usd
            a["musteriler"].add(c); a["fatura"] += 1
            a["Ulke_EN"] = f["ulke_en"] or country_geo.get(u,{}).get("Ulke_EN", u)

        # Müşteri
        m = cari_agg[c]
        m["ad"] = f["unvan"] or (str(c).zfill(8) if isinstance(c,int) else str(c))
        m["ulke"] = u
        m["toplam"] += usd; m["fatura"] += 1
        if f["dt"]:
            if m["son"] is None or f["dt"] > m["son"]: m["son"] = f["dt"]
            if m["ilk"] is None or f["dt"] < m["ilk"]: m["ilk"] = f["dt"]

        # Ürün
        p = urun_agg[g]
        p["tip"] = f["tip"]; p["seri"] = f["seri"]
        p["miktar"] += mik; p["usd"] += usd
        if yil in YEARS: p[f"usd_{yil}"] += usd
        p["renkler"][f["renk"]]["usd"] += usd
        p["renkler"][f["renk"]]["miktar"] += mik

        # Renk (ürün bazında)
        rk = (g, f["renk"])
        renk_agg[rk]["USD_Tutar"] += usd
        renk_agg[rk]["miktar"] += mik

        # Kategori (UrunTipi)
        if f["tip"]:
            kat_agg[f["tip"]]["USD_Tutar"] += usd
            kat_agg[f["tip"]]["miktar"] += mik

        # Trend
        if yil: trend_y[yil] += usd
        if yil and f["ay"]: trend_m[(yil, f["ay"])] += usd

    # ---------- ULKE listesi ----------
    ulke_list = []
    for u, a in ulke_agg.items():
        geo = country_geo.get(u, {})
        buyume = round(((a["usd_2025"]-a["usd_2024"])/a["usd_2024"]*100),1) if a["usd_2024"]>0 else 0.0
        ulke_list.append({
            "Ulke":u, "miktar":round(a["miktar"]),
            "usd_2023":round(a["usd_2023"],2),"usd_2024":round(a["usd_2024"],2),
            "usd_2025":round(a["usd_2025"],2),"usd_2026":round(a["usd_2026"],2),
            "musteri":len(a["musteriler"]), "fatura":a["fatura"],
            "Ulke_EN":a["Ulke_EN"] or geo.get("Ulke_EN",u),
            "buyume":buyume,
            "lat":geo.get("lat",0), "lon":geo.get("lon",0),
            "kg":0, "usd_kgcov":0,
            "toplam":round(a["toplam"],2),
        })
    ulke_list.sort(key=lambda x:-x["toplam"])

    # ---------- MÜŞTERİ listesi + RFM ----------
    mus_list = []
    for c, m in cari_agg.items():
        gun = (max_dt - m["son"]).days if m["son"] else 9999
        seg = rfm_segment(gun, m["toplam"], m["fatura"])
        mus_list.append({
            "kod": str(c).zfill(8) if isinstance(c,int) else str(c),
            "ad": m["ad"], "ulke": m["ulke"],
            "toplam": round(m["toplam"],2), "fatura": m["fatura"],
            "son_tarih": m["son"].strftime("%Y-%m-%d") if m["son"] else "",
            "ilk_tarih": m["ilk"].strftime("%Y-%m-%d") if m["ilk"] else "",
            "gun": gun, "seg": seg,
        })
    mus_list.sort(key=lambda x:-x["toplam"])

    # ---------- FİRMALAR (koordinatı olan müşteriler) ----------
    firmalar = []
    for m in mus_list:
        geo = country_geo.get(m["ulke"])
        if not geo or not geo.get("lat"): continue
        firmalar.append({
            "ad":m["ad"], "ulke":m["ulke"],
            "lat":geo["lat"], "lon":geo["lon"],
            "toplam":m["toplam"], "fatura":m["fatura"],
            "ilk":m["ilk_tarih"], "son":m["son_tarih"],
        })

    # ---------- ÜRÜNLER ----------
    toplam_usd = sum(p["usd"] for p in urun_agg.values()) or 1
    urunler = []
    for g, p in urun_agg.items():
        renkler = [{"Renk":rk,"usd":round(rv["usd"],2),"miktar":round(rv["miktar"])}
                   for rk,rv in p["renkler"].items()]
        renkler.sort(key=lambda x:-x["miktar"])   # QTY büyükten küçüğe
        urunler.append({
            "ad":g, "ad_en":prod_en.get(g, g),
            "tip":p["tip"], "seri":p["seri"],
            "seri_cat":seri_cat.get(p["seri"], p["seri"]),
            "miktar":round(p["miktar"]), "usd":round(p["usd"],2),
            "pay":round(p["usd"]/toplam_usd*100,2),
            "usd_2023":round(p["usd_2023"],2),"usd_2024":round(p["usd_2024"],2),
            "usd_2025":round(p["usd_2025"],2),"usd_2026":round(p["usd_2026"],2),
            "renkler":renkler,
        })
    urunler.sort(key=lambda x:-x["miktar"])

    # ---------- RENK ANALİZ ----------
    renk_analiz = [{"GenelUrunAdi":g,"Renk":rk,
                    "USD_Tutar":round(v["USD_Tutar"],2),"miktar":round(v["miktar"])}
                   for (g,rk),v in renk_agg.items()]
    renk_analiz.sort(key=lambda x:-x["USD_Tutar"])

    # ---------- KATEGORİ ----------
    kategori = [{"Kategori":t,"USD_Tutar":round(v["USD_Tutar"],2),"miktar":round(v["miktar"])}
                for t,v in kat_agg.items()]
    kategori.sort(key=lambda x:-x["USD_Tutar"])

    # ---------- TREND ----------
    trend = [{"Yil":y,"USD":round(trend_y[y],2)} for y in sorted(trend_y)]
    AYLAR = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
    trend_monthly = [{"label":f"{AYLAR[ay-1]} {yil}","yil":yil,"ay":ay,"USD":round(v,2)}
                     for (yil,ay),v in sorted(trend_m.items())]

    # ---------- KPI ----------
    y = {yy: round(trend_y.get(yy,0.0),2) for yy in YEARS}
    total4 = round(sum(y.values()),2)
    p2025 = round((y[2025]-y[2024])/y[2024]*100,1) if y[2024]>0 else 0.0
    p2024 = round((y[2024]-y[2023])/y[2023]*100,1) if y[2023]>0 else 0.0
    # son 12 ay (referans tarihten geriye)
    son12 = 0.0
    cutoff = (max_dt.year*12 + max_dt.month) - 12
    for (yil,ay),v in trend_m.items():
        if yil*12+ay > cutoff: son12 += v
    kpi = {
        "total4":total4,
        "y2025":y[2025],"usd_2025":y[2025],
        "y2024":y[2024],"usd_2024":y[2024],
        "y2023":y[2023],"usd_2023":y[2023],
        "ytd2026":y[2026],"usd_2026":y[2026],
        "p2025":p2025,"p2024":p2024,
        "ulke":len(ulke_list),"musteri":len(mus_list),
        "ulke_sayisi":len(ulke_list),"musteri_sayisi":len(mus_list),
        "son12":round(son12,2),
    }

    report["totals"] = {f"y{yy}":y[yy] for yy in YEARS}
    report["counts"] = {"ulke":len(ulke_list),"musteri":len(mus_list),
                        "firma":len(firmalar),"urun":len(urunler)}

    return {
        "ulke":ulke_list, "kpi":kpi, "musteri_listesi":mus_list,
        "firmalar":firmalar, "renk_analiz":renk_analiz,
        "trend":trend, "trend_monthly":trend_monthly,
        "urunler":urunler, "kategori":kategori,
    }

def rfm_segment(gun, toplam, fatura):
    """
    RFM segmentasyonu — recency (gün) + monetary (USD) kombinasyonu.
    Eşikler yapılandırılabilir; dashboard efsanesiyle uyumlu isimler üretir.
    """
    if gun >= 730:              # ~2 yıldır alım yok
        return "lost"
    if gun >= 354:              # ~1-2 yıl
        return "risk"
    # son ~1 yıl içinde alım yapmış -> değere göre
    if toplam >= 150000 and gun <= 150:
        return "star"
    if toplam >= 100000:
        return "loyal"
    if toplam >= 20000 and gun <= 130:
        return "star"
    if toplam >= 20000:
        return "loyal"
    return "potential"

# ══════════════════════════════════════════════════════════════════
#  4) HTML ENJEKSİYONU
# ══════════════════════════════════════════════════════════════════
def inject_D(html_path, D, out_path):
    with open(html_path, encoding="utf-8") as f:
        html = f.read()
    anchor = "const D = "
    i = html.find(anchor)
    if i < 0:
        raise RuntimeError("dashboard.html içinde 'const D = ' bulunamadı")
    j = html.find("{", i)
    depth = 0; end = j
    for k, c in enumerate(html[j:]):
        if c == "{": depth += 1
        elif c == "}":
            depth -= 1
            if depth == 0:
                end = j + k + 1
                break
    new_json = json.dumps(D, ensure_ascii=False, separators=(",", ":"))
    new_html = html[:j] + new_json + html[end:]
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(new_html)
    return len(new_json)

# ══════════════════════════════════════════════════════════════════
#  5) ÇALIŞTIR
# ══════════════════════════════════════════════════════════════════
def main():
    ap = argparse.ArgumentParser(description="SmartPack dashboard veri derleyici")
    ap.add_argument("--ham", default=PATH_HAM)
    ap.add_argument("--tax", default=PATH_TAX)
    ap.add_argument("--html", default=PATH_HTML)
    ap.add_argument("--out", default=PATH_HTML)
    ap.add_argument("--check", action="store_true", help="sadece doğrula, dosya yazma")
    args = ap.parse_args()

    print("═"*64)
    print("  SmartPack Export Dashboard — DATA AGENT")
    print("═"*64)

    ref = {
        "product_en":   load_json(os.path.join(REF,"product_en.json")),
        "seri_cat":     load_json(os.path.join(REF,"seri_cat.json")),
        "country_geo":  load_json(os.path.join(REF,"country_geo.json")),
        "cari_country": load_json(os.path.join(REF,"cari_country.json")),
        "exchange_rates": load_json(os.path.join(REF,"exchange_rates.json"),
                                     {"USD":1.0,"EUR":1.08,"GBP":1.27,"TRY":0.031}),
    }
    rates = ref["exchange_rates"]
    report = {}

    print(f"► Taksonomi okunuyor: {os.path.basename(args.tax)}")
    tax = load_taxonomy(args.tax)
    print(f"  {len(tax)} stok kodu yüklendi")

    facts = []
    # 1) Ham DIA fatura dosyaları (varsa öncelikli), yoksa hazır ham_veri tablosu
    if os.path.isdir(DIR_DIA) and any(f.lower().endswith((".xlsx",".xls"))
                                       for f in os.listdir(DIR_DIA)):
        print(f"► Ham DIA faturaları okunuyor: data/dia/")
        dia_facts = load_facts_from_dia(DIR_DIA, tax, ref["cari_country"], rates, report)
        print(f"  {len(dia_facts)} direkt ihracat satırı ({len(report.get('dia_files',[]))} dosya)")
        facts += dia_facts
    else:
        print(f"► Ham fatura okunuyor: {os.path.basename(args.ham)}")
        facts += load_facts(args.ham, tax, rates, report)
        print(f"  {len(facts)} direkt satır")

    # 2) İhraç kayıtlı (indirekt) satışlar
    if os.path.exists(PATH_IND):
        print(f"► İhraç kayıtlı satışlar okunuyor: {os.path.basename(PATH_IND)}")
        ind_facts = load_facts_from_indirect(PATH_IND, tax, rates, report)
        print(f"  {len(ind_facts)} ihraç kayıtlı satır")
        facts += ind_facts

    print(f"► Toplam {len(facts)} geçerli fatura satırı")

    print("► Veri derleniyor...")
    D = build_D(facts, tax, ref, report)

    # -------- DOĞRULAMA RAPORU --------
    print("\n" + "─"*64)
    print("  DOĞRULAMA RAPORU")
    print("─"*64)
    t = report["totals"]
    print(f"  Yıllık ciro (USD):")
    for yy in YEARS:
        print(f"     {yy}: ${t[f'y{yy}']:>14,.2f}")
    print(f"     TOPLAM: ${sum(t.values()):>11,.2f}")
    c = report["counts"]
    print(f"  Ülke: {c['ulke']}  |  Müşteri: {c['musteri']}  |  "
          f"Firma(haritada): {c['firma']}  |  Ürün: {c['urun']}")

    uc = report.get("unknown_codes", {})
    if uc:
        print(f"\n  ⚠ TAKSONOMİDE OLMAYAN {len(uc)} STOK KODU "
              f"({sum(uc.values())} satır atlandı):")
        for code, n in sorted(uc.items(), key=lambda x:-x[1])[:20]:
            print(f"     StokKodu {code}: {n} satır  → taksonomi.xlsx'e ekleyin")
    else:
        print("\n  ✓ Tüm stok kodları taksonomide eşleşti")

    # eksik EN çeviri / koordinat uyarıları
    missing_en = [u["ad"] for u in D["urunler"] if u["ad_en"]==u["ad"]]
    if missing_en:
        print(f"\n  ⚠ {len(missing_en)} ürünün İngilizce adı eksik "
              f"(product_en.json): {missing_en[:5]}")
    missing_geo = sorted({u["Ulke"] for u in D["ulke"]
                          if not u["lat"] and u["Ulke"] != IND_MARKET})
    if missing_geo:
        print(f"  ⚠ Koordinatı eksik ülke(ler): {missing_geo} "
              f"→ country_geo.json'a ekleyin")

    if args.check:
        print("\n[--check] Doğrulama modu — dosya yazılmadı.")
        return

    print("\n► dashboard.html güncelleniyor...")
    size = inject_D(args.html, D, args.out)
    print(f"  ✓ D enjekte edildi ({size:,} bayt) → {args.out}")
    print("\n✅ Tamamlandı.")

if __name__ == "__main__":
    main()
